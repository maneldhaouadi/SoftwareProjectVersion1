from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from .models import Employe
import os
from django.http import HttpResponse
from django.template.loader import render_to_string
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# DÉCORATEUR DE PROTECTION
def protected_view(view_func):
    def wrapper(request, *args, **kwargs):
        if 'employe_id' not in request.session:
            messages.error(request, "Vous devez vous connecter.")
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper

# LOGIN (inchangé, correct)
def login_view(request):
    if request.method == 'POST':
        login_input = request.POST.get('login', '').strip()
        password = request.POST.get('password', '')

        if not login_input or not password:
            messages.error(request, "Veuillez remplir tous les champs.")
            return render(request, 'employes/login.html')

        try:
            employe = Employe.objects.get(login=login_input)
        except Employe.DoesNotExist:
            messages.error(request, "Login introuvable.")
            return render(request, 'employes/login.html')

        if employe.check_password(password):
            request.session['employe_id'] = employe.id
            request.session['employe_nom'] = f"{employe.prenom} {employe.nom}"
            request.session['employe_role'] = employe.get_role_display()
            request.session['employe_login'] = employe.login
            messages.success(request, f"Bienvenue, {employe.prenom} !")
            return redirect('liste_employes')
        else:
            messages.error(request, "Mot de passe incorrect.")
    return render(request, 'employes/login.html')

def logout_view(request):
    request.session.flush()
    messages.success(request, "Vous êtes déconnecté.")
    return redirect('login')

#LISTE 
@protected_view
def liste_employes(request):
    if request.session.get('employe_role') != 'Administratif':
        messages.error(request, "Accès refusé.")
        return redirect('login')

    employes = Employe.objects.all()

    # Recherche
    if request.GET.get('q'):
        q = request.GET['q'].strip()
        employes = employes.filter(
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(email__icontains=q) |
            Q(login__icontains=q)
        )

    # Filtres
    if request.GET.get('role'):
        employes = employes.filter(role=request.GET['role'])
    if request.GET.get('etat'):
        employes = employes.filter(etat=request.GET['etat'])

    # TRI
    tri = request.GET.get('tri', 'date_ajout_desc')
    if tri == 'nom_asc':
        employes = employes.order_by('nom', 'prenom')
    elif tri == 'nom_desc':
        employes = employes.order_by('-nom', '-prenom')
    elif tri == 'date_embauche_asc':
        employes = employes.order_by('dateEmbauche')
    elif tri == 'date_embauche_desc':
        employes = employes.order_by('-dateEmbauche')
    else:  # date_ajout_desc (par défaut)
        employes = employes.order_by('-id')  # id décroissant = plus récent en premier

    context = {
        'employes': employes,
        'total': employes.count(),
        'roles': Employe.ROLE_CHOICES,
        'etats': Employe.ETAT_CHOICES,
    }
    return render(request, 'employes/liste.html', context)
#AJOUTER EMPLOYÉ

@protected_view
def ajouter_employe(request):
    if request.session.get('employe_role') != 'Administratif':
        messages.error(request, "Accès refusé.")
        return redirect('login')

    # On passe toujours les choix au template (GET + POST)
    context = {
        'ROLE_CHOICES': Employe.ROLE_CHOICES,
        'ETAT_CHOICES': Employe.ETAT_CHOICES,
    }

    if request.method == 'POST':
        try:
            mot_de_passe = request.POST.get('motDePasse', '')
            if not mot_de_passe:
                messages.error(request, "Le mot de passe est obligatoire.")
                return render(request, 'employes/ajouter.html', context)

            employe = Employe(
                nom=request.POST['nom'].strip(),
                prenom=request.POST['prenom'].strip(),
                role=request.POST['role'],
                login=request.POST['login'].strip(),
                email=request.POST['email'].strip(),
                telephone=request.POST.get('telephone', '') or None,
                dateEmbauche=request.POST['dateEmbauche'],
                etat=request.POST.get('etat', 'ACTIF'),
            )

            if 'photo' in request.FILES:
                employe.photo = request.FILES['photo']

            employe.set_password(mot_de_passe)
            employe.save()

            messages.success(request, f"Employé {employe.prenom} {employe.nom} ajouté avec succès !")
            return redirect('liste_employes')

        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")
            return render(request, 'employes/ajouter.html', context)  # on repasse les choix même en cas d'erreur

    return render(request, 'employes/ajouter.html', context)  # GET normal


# ====================== MODIFIER EMPLOYÉ (CORRIGÉ) ======================
@protected_view
def modifier_employe(request, id):
    if request.session.get('employe_role') != 'Administratif':
        messages.error(request, "Accès refusé.")
        return redirect('login')

    employe = get_object_or_404(Employe, id=id)

    # Contexte de base (toujours envoyé)
    context = {
        'employe': employe,
        'ROLE_CHOICES': Employe.ROLE_CHOICES,
        'ETAT_CHOICES': Employe.ETAT_CHOICES,
    }

    if request.method == 'POST':
        try:
            employe.nom = request.POST['nom'].strip()
            employe.prenom = request.POST['prenom'].strip()
            employe.role = request.POST['role']
            employe.login = request.POST['login'].strip()
            employe.email = request.POST['email'].strip()
            employe.telephone = request.POST.get('telephone', '') or None
            employe.dateEmbauche = request.POST['dateEmbauche']
            employe.etat = request.POST.get('etat', 'ACTIF')

            if 'photo' in request.FILES:
                if employe.photo and os.path.isfile(employe.photo.path):
                    os.remove(employe.photo.path)
                employe.photo = request.FILES['photo']

            nouveau_mdp = request.POST.get('motDePasse', '')
            if nouveau_mdp:
                employe.set_password(nouveau_mdp)

            employe.save()
            messages.success(request, "Employé modifié avec succès !")
            return redirect('liste_employes')

        except Exception as e:
            messages.error(request, f"Erreur : {str(e)}")

    return render(request, 'employes/modifier.html', context)

# ====================== SUPPRIMER ======================
@protected_view
def supprimer_employe(request, id):
    if request.session.get('employe_role') != 'Administratif':
        messages.error(request, "Accès refusé.")
        return redirect('login')

    employe = get_object_or_404(Employe, id=id)
    if request.method == 'POST':
        # Supprimer la photo du disque
        if employe.photo and os.path.isfile(employe.photo.path):
            os.remove(employe.photo.path)
        employe.delete()
        messages.success(request, "Employé supprimé définitivement.")
        return redirect('liste_employes')
    return render(request, 'employes/supprimer.html', {'employe': employe})

# ====================== DÉTAIL ======================
@protected_view
def detail_employe(request, id):
    if request.session.get('employe_role') != 'Administratif':
        messages.error(request, "Accès refusé.")
        return redirect('login')
    employe = get_object_or_404(Employe, id=id)
    return render(request, 'employes/detail.html', {'employe': employe})
# ====================== EXPORT PDF ======================
# ====================== EXPORT PDF (100% FONCTIONNEL SUR WINDOWS) ======================
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from datetime import date

@protected_view
def export_pdf_employes(request):
    if request.session.get('employe_role') != 'Administratif':
        messages.error(request, "Accès refusé.")
        return redirect('login')

    # Même logique que liste_employes
    employes = Employe.objects.all()

    if request.GET.get('q'):
        q = request.GET['q'].strip()
        employes = employes.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) |
            Q(email__icontains=q) | Q(login__icontains=q) | Q(cin__icontains=q)
        )
    if request.GET.get('role'):
        employes = employes.filter(role=request.GET['role'])
    if request.GET.get('etat'):
        employes = employes.filter(etat=request.GET['etat'])

    tri = request.GET.get('tri', 'date_ajout_desc')
    if tri == 'nom_asc':
        employes = employes.order_by('nom', 'prenom')
    elif tri == 'nom_desc':
        employes = employes.order_by('-nom', '-prenom')
    elif tri == 'date_embauche_asc':
        employes = employes.order_by('dateEmbauche')
    elif tri == 'date_embauche_desc':
        employes = employes.order_by('-dateEmbauche')
    else:
        employes = employes.order_by('-id')

    context = {
        'employes': employes,
        'total': employes.count(),
        'filtres': request.GET.dict(),
        'date_generation': date.today().strftime("%d/%m/%Y"),
    }

    html = render_to_string('employes/pdf_employes.html', context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="employes_{date.today().strftime("%Y%m%d")}.pdf"'

    pdf = pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=response)
    if not pdf.err:
        return response
    return HttpResponse("Erreur génération PDF", status=400)
# ====================== EXPORT EXCEL ======================
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

@protected_view
def export_excel_employes(request):
    if request.session.get('employe_role') != 'Administratif':
        messages.error(request, "Accès refusé.")
        return redirect('login')

    # Même logique que liste_employes
    employes = Employe.objects.all()

    if request.GET.get('q'):
        q = request.GET['q'].strip()
        employes = employes.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) |
            Q(email__icontains=q) | Q(login__icontains=q) | Q(cin__icontains=q)
        )
    if request.GET.get('role'):
        employes = employes.filter(role=request.GET['role'])
    if request.GET.get('etat'):
        employes = employes.filter(etat=request.GET['etat'])

    tri = request.GET.get('tri', 'date_ajout_desc')
    if tri == 'nom_asc':
        employes = employes.order_by('nom', 'prenom')
    elif tri == 'nom_desc':
        employes = employes.order_by('-nom', '-prenom')
    elif tri == 'date_embauche_asc':
        employes = employes.order_by('dateEmbauche')
    elif tri == 'date_embauche_desc':
        employes = employes.order_by('-dateEmbauche')
    else:
        employes = employes.order_by('-id')

    # Création du fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Employés"

    # En-tête
    columns = ['Nom Prénom', 'Rôle', 'État', 'Email', 'Téléphone', 'Embauche', 'Ajouté le']
    ws.append(columns)

    # Style en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4158d0", end_color="4158d0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Données
    for e in employes:
        ws.append([
            f"{e.prenom} {e.nom}",
            e.get_role_display(),
            e.get_etat_display(),
            e.email,
            e.telephone or "",
            e.dateEmbauche.strftime("%d/%m/%Y"),
            e.id  # ou e.date_ajout si tu as un champ created_at
        ])

    # Ajuster la largeur des colonnes
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="employes_{date.today().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response
# ====================== DASHBOARD ======================
@protected_view
def dashboard(request):
    if request.session.get('employe_role') != 'Administratif':
        messages.error(request, "Accès refusé.")
        return redirect('login')

    total_employes = Employe.objects.count()
    actifs = Employe.objects.filter(etat='ACTIF').count()
    inactifs = Employe.objects.filter(etat='INACTIF').count()
    conge = Employe.objects.filter(etat='CONGE').count()
    suspension = Employe.objects.filter(etat='SUSPENSION').count()

    # Statistiques par rôle
    roles_stats = []
    for role_code, role_name in Employe.ROLE_CHOICES:
        count = Employe.objects.filter(role=role_code).count()
        roles_stats.append({'role': role_name, 'count': count, 'percent': round(count/total_employes*100, 1) if total_employes else 0})

    # Employés par mois d'embauche (dernières 12 mois)
    from django.utils import timezone
    from datetime import datetime
    from collections import defaultdict
    import calendar

    embauches_par_mois = defaultdict(int)
    for e in Employe.objects.all():
        mois = e.dateEmbauche.strftime("%Y-%m")
        embauches_par_mois[mois] += 1

    # Derniers 12 mois
    mois_labels = []
    mois_valeurs = []
    today = timezone.now()
    for i in range(11, -1, -1):
        mois_date = today.replace(day=1) - timezone.timedelta(days=30*i)
        mois_key = mois_date.strftime("%Y-%m")
        mois_label = calendar.month_name[mois_date.month][:3] + " " + str(mois_date.year)
        mois_labels.append(mois_label)
        mois_valeurs.append(embauches_par_mois.get(mois_key, 0))

    context = {
        'total_employes': total_employes,
        'actifs': actifs,
        'inactifs': inactifs,
        'conge': conge,
        'suspension': suspension,
        'roles_stats': roles_stats,
        'mois_labels': mois_labels,
        'mois_valeurs': mois_valeurs,
        'pourcentage_actifs': round(actifs/total_employes*100, 1) if total_employes else 0,
    }
    return render(request, 'employes/dashboard.html', context)
def accueil(request):
    return render(request, 'index.html')

def apropos(request):
    return render(request, 'about.html')

def contact(request):
    return render(request, 'contact.html')